import pandas as pd
import openpyxl
from collections import Counter
import re
import json

file_path = 'decrypted_sample.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

# 수식 분석 함수
def analyze_formulas_detailed(ws, sheet_name, max_samples=50):
    formulas = []
    formula_types = Counter()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': formula[:100] + ('...' if len(formula) > 100 else '')  # 길이 제한
                })
                
                # 함수 타입 분류
                if 'SUMIFS' in formula:
                    formula_types['SUMIFS'] += 1
                elif 'VLOOKUP' in formula:
                    formula_types['VLOOKUP'] += 1
                elif 'IF(' in formula:
                    formula_types['IF'] += 1
                elif 'SUM(' in formula:
                    formula_types['SUM'] += 1
                elif 'SUBTOTAL' in formula:
                    formula_types['SUBTOTAL'] += 1
                elif '+' in formula or '-' in formula or '*' in formula or '/' in formula:
                    formula_types['ARITHMETIC'] += 1
                else:
                    formula_types['OTHER'] += 1
    
    return formulas, formula_types

print('=== 수식 상세 분석 ===')

# 수식이 많은 주요 시트들 분석
important_sheets = ['출', '월별요약손익계산서(추정)', '손익계산서추이분석', '경영분석']

all_formula_stats = {}

for sheet_name in important_sheets:
    if sheet_name in wb.sheetnames:
        print(f'\n--- {sheet_name} 시트 ---')
        ws = wb[sheet_name]
        formulas, formula_types = analyze_formulas_detailed(ws, sheet_name)
        
        all_formula_stats[sheet_name] = {
            'total_formulas': len(formulas),
            'formula_types': dict(formula_types)
        }
        
        print(f'총 수식 개수: {len(formulas)}개')
        print('수식 타입별 분포:')
        for func_type, count in formula_types.most_common():
            print(f'  {func_type}: {count}개')
        
        # 샘플 수식들 표시
        print('수식 샘플 (처음 10개):')
        for i, formula_info in enumerate(formulas[:10]):
            print(f'  {formula_info["cell"]}: {formula_info["formula"]}')

# 월별요약손익계산서의 구조 분석
print('\n=== 월별요약손익계산서 구조 분석 ===')
if '월별요약손익계산서(추정)' in wb.sheetnames:
    ws_monthly = wb['월별요약손익계산서(추정)']
    
    # A열의 항목들 (손익계산서 항목들)
    print('손익계산서 항목들:')
    for row_idx in range(1, min(30, ws_monthly.max_row + 1)):
        item = ws_monthly.cell(row=row_idx, column=1).value
        if item and not str(item).isdigit():
            print(f'  행{row_idx}: {item}')

print('\n전체 수식 통계 요약:')
print(json.dumps(all_formula_stats, indent=2, ensure_ascii=False))