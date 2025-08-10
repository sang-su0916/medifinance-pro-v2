import pandas as pd
import openpyxl
from collections import Counter
import json

file_path = 'decrypted_sample.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print('=== 매출내역total 시트 분석 ===')
ws_sales = wb['매출내역total']

# 헤더 분석
print('매출내역total 시트 구조:')
for row_idx in range(1, 6):
    print(f'\n{row_idx}행:')
    row_data = []
    for col_idx in range(1, ws_sales.max_column + 1):
        cell_value = ws_sales.cell(row=row_idx, column=col_idx).value
        if cell_value is not None:
            row_data.append(f'{col_idx}열: {cell_value}')
    
    if row_data:
        print('  ' + ', '.join(row_data[:8]))  # 처음 8개만 표시

print('\n=== 사업장요약현황 시트 분석 ===')
ws_summary = wb['사업장요약현황']

print('사업장요약현황 시트 구조:')
for row_idx in range(1, min(20, ws_summary.max_row + 1)):
    # A열과 O열 위주로 분석 (계정과목 관련)
    a_value = ws_summary.cell(row=row_idx, column=1).value
    o_value = ws_summary.cell(row=row_idx, column=15).value  # O열
    
    if a_value or o_value:
        print(f'  행{row_idx}: A열={a_value}, O열={o_value}')

print('\n=== 시트간 참조 관계 분석 ===')
wb_with_formulas = openpyxl.load_workbook(file_path, data_only=False)

# 주요 시트들 간의 참조 관계
sheet_references = {}

for sheet_name in ['출', '분', '매출내역total', '월별요약손익계산서(추정)', '사업장요약현황']:
    ws = wb_with_formulas[sheet_name]
    references = set()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                # 다른 시트 참조 찾기
                for other_sheet in wb.sheetnames:
                    if other_sheet != sheet_name and other_sheet in formula:
                        references.add(other_sheet)
    
    sheet_references[sheet_name] = list(references)

print('시트간 참조 관계:')
for sheet, refs in sheet_references.items():
    if refs:
        print(f'  {sheet} → {", ".join(refs)}')

print('\n=== 월별요약손익계산서 세부 구조 ===')
ws_monthly = wb['월별요약손익계산서(추정)']

# B열의 계정과목들 분석
print('손익계산서 계정 항목들:')
for row_idx in range(1, min(50, ws_monthly.max_row + 1)):
    b_value = ws_monthly.cell(row=row_idx, column=2).value
    if b_value and not str(b_value).isdigit() and b_value != '계정과목':
        print(f'  행{row_idx}: {b_value}')

print('\n=== 입력/계산/출력 시트 분류 ===')
sheet_classification = {
    '입력시트': ['출', '분', '매출내역total'],
    '계산시트': ['월별요약손익계산서(추정)', '손익계산서추이분석', '경영분석'],
    '요약/출력시트': ['사업장요약현황'],
    '기타': ['추가확인6.30', '표지', '작성시유의사항']
}

for category, sheets in sheet_classification.items():
    print(f'\n{category}:')
    for sheet in sheets:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f'  {sheet}: {ws.max_row-1}행 데이터, {ws.max_column}열')

print('\n=== 분석 완료 ===')