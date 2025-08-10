import pandas as pd
import openpyxl
from collections import Counter
import json

file_path = 'decrypted_sample.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print('=== 계정과목 체계 분석 ===')

# 출 시트에서 계정과목 분석 (10열)
ws_out = wb['출']
account_subjects_out = []
for row_idx in range(4, ws_out.max_row + 1):
    account = ws_out.cell(row=row_idx, column=10).value
    if account and str(account) != '계정과목':
        account_subjects_out.append(str(account))

# 분 시트에서 계정과목 분석 (10열)  
ws_bun = wb['분']
account_subjects_bun = []
for row_idx in range(2, ws_bun.max_row + 1):
    account = ws_bun.cell(row=row_idx, column=10).value
    if account and str(account) != '계정과목':
        account_subjects_bun.append(str(account))

print('출 시트 계정과목 분포:')
out_counter = Counter(account_subjects_out)
for account, count in out_counter.most_common(20):
    print(f'  {account}: {count}건')

print('\n분 시트 계정과목 분포:')
bun_counter = Counter(account_subjects_bun)
for account, count in bun_counter.most_common(20):
    print(f'  {account}: {count}건')

print(f'\n출 시트 총 거래 건수: {len(account_subjects_out)}건')
print(f'분 시트 총 거래 건수: {len(account_subjects_bun)}건')
print(f'출 시트 계정과목 종류: {len(set(account_subjects_out))}개')
print(f'분 시트 계정과목 종류: {len(set(account_subjects_bun))}개')

# 대분류 분석 (출 시트 9열)
print('\n=== 대분류 체계 분석 (출 시트) ===')
large_categories = []
for row_idx in range(4, ws_out.max_row + 1):
    category = ws_out.cell(row=row_idx, column=9).value
    if category:
        large_categories.append(str(category))

large_counter = Counter(large_categories)
for category, count in large_counter.most_common():
    print(f'  {category}: {count}건')

# 소분류 분석 (출 시트 8열)
print('\n=== 소분류 체계 분석 (출 시트) ===')
small_categories = []
for row_idx in range(4, ws_out.max_row + 1):
    category = ws_out.cell(row=row_idx, column=8).value
    if category:
        small_categories.append(str(category))

small_counter = Counter(small_categories)
for category, count in small_counter.most_common(15):  # 상위 15개
    print(f'  {category}: {count}건')